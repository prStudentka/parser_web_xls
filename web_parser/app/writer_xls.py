import os
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from typing import List


__TITLE:     str = 'tutorials'
__FILE_NAME: str = 'Habr tutorials'
__SUFFIX:    str = '.xlsx'
__DIRECTORY: str = 'documents'
__WIDTH_COLUMN: int = 70
__TITLES_ROW: List[str] = ['title tutorials',
                           'url tutorials']


def __decorated_column(work_sheet: Worksheet) -> None:
    '''Decorated columns worksheet

    :param work_book: Workbook

    Function changed font and
    dimensions for column in the worksheet.
    '''

    font: Font = Font(name='Arial', size=12, bold=True)
    border: Border = Border(bottom=Side(border_style='thin',
                                        color='000000'))
    work_sheet.append(__TITLES_ROW)
    for i, _ in enumerate(__TITLES_ROW, 1):
        letter: str = get_column_letter(i)
        work_sheet[f'{letter}1'].font = font
        work_sheet[f'{letter}1'].alignment = Alignment(horizontal="center",
                                                       wrap_text=True)
        work_sheet[f'{letter}1'].border = border
        work_sheet.column_dimensions[f'{letter}'].width = __WIDTH_COLUMN
        work_sheet.column_dimensions[letter].auto_size = True


def __changed_directory() -> int:
    '''Change working dirrectory

    :param page: int - page number

    Function add directory for save. If directory
    not exist - make it. Then change directory and return
    1 if all OK or
    0 if error.
    '''

    path_directory = os.path.join(os.getcwd(), __DIRECTORY)
    try:
        os.makedirs(path_directory, exist_ok=True)
    except OSError:
        raise ('No write permission in the root file system')
        return 0
    os.chdir(path=path_directory)
    return 1


def save_xls(data: list) -> tuple:
    '''Saved in xls data

    :param data: list - rows for Excel

    Function create workbook and add rows in the worksheet.
    Save it in your directory. Return message about process.
    Ok - if all right
    Bad - if not good
    '''

    if not data:
        raise ('Not found informations')
        return ('Bad',)
    work_book: Workbook = Workbook()
    work_sheet: Worksheet = work_book.active
    work_sheet = work_book.create_sheet(title=__TITLE, index=0)
    __decorated_column(work_sheet)
    for row in data:
        work_sheet.append(row)
    if __changed_directory():
        name: str = f'{__FILE_NAME}_{date.today()}{__SUFFIX}'
        work_book.save(name)
        work_book.close()
        return ('OK', os.path.join(os.getcwd(), name))
    return ('Bad', data)
